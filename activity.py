import requests 
import sched, time
import pytz
from pytz import timezone
utc = pytz.utc
eastern = timezone('America/Montreal')
from datetime import datetime

from openpyxl import Workbook, load_workbook
from ask_schools import school_name, find_school_by_operator_suffix

fmt_date = '%Y-%m-%d'
fmt_hour = '%H:%M:%S'
fmt_hour_round = '%H.%M'
from pprint import pprint as print
import os

import lh3.api
client = lh3.api.Client()
client.set_options(version = 'v1')
# For each user... 

def prepare_workbook():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'date'
    ws['B1'] = 'time'
    ws['C1'] = 'time_hour'
    ws['D1'] = 'operator'
    ws['E1'] = 'operator_status'
    ws['F1'] = 'school'
    ws['G1'] = 'scholars-portal'
    ws['H1'] = 'scholars-portal-txt'
    ws['I1'] = 'clavardez'
    ws['J1'] = 'practice-webinars'
    ws['K1'] = 'practice-webinars-fr'
    ws['L1'] = 'toronto-mississauga'
    ws['M1'] = 'toronto-scarborough'
    ws['N1'] = 'toronto-st-george'
    ws['O1'] = 'brock'
    ws['P1'] = 'carleton'
    ws['Q1'] = 'carleton-txt'
    ws['R1'] = 'laurentian'
    ws['S1'] = 'laurentian-fr'
    ws['T1'] = 'otech'
    ws['U1'] = 'queens'
    ws['V1'] = 'western'
    ws['W1'] = 'western-txt'
    ws['X1'] = 'western-proactive'

    loc_dt = eastern.localize(datetime.now())
    if os.environ['environment'] == 'prod':
        wb.save("/root/sp_lh3_monitoring_script/" +str(loc_dt.strftime(fmt_date))+"-activity" +'.xlsx')
    else:
        wb.save(str(loc_dt.strftime(fmt_date))+"-activity" +'.xlsx')
    return wb
    
def get_filename():
    loc_dt = eastern.localize(datetime.now())
    if os.environ['environment'] == 'prod':
        filename = "/root/sp_lh3_monitoring_script/" +str(loc_dt.strftime(fmt_date)) +"-activity" +'.xlsx'
    else:
        filename = str(loc_dt.strftime(fmt_date)) +"-activity" +'.xlsx'
    return filename

def try_open_file():
    filename = get_filename()
    try:
        wb = load_workbook(filename)
        return wb
    except:
        wb = prepare_workbook()
        return wb

def insert_to_excel(availabilities, user):
    wb = try_open_file()
    ws = wb.active
    max_row = ws.max_row + 1
    for value in availabilities:
        ws['A'+str(max_row)] = value.get('date')
        ws['B'+str(max_row)] = value.get('time')
        ws['C'+str(max_row)] = value.get('hour_floor')
        ws['D'+str(max_row)] = value.get('operator')

        if user.get('status') == 'null':
            ws['E'+str(max_row)] = value.get('user_show')
            if  value.get('user_show') in  ['dnd', 'away', 'xa']:
                ws['E'+str(max_row)] = 'unavailable-'+ value.get('user_show')
            else:
                ws['E'+str(max_row)] = 'available'
        else:
            ws['E'+str(max_row)] = user.get('status')

        user_show_value = value.get('user_show') 
        if user_show_value == 'chat':
            user_show_value = 'available'

        ws['F'+str(max_row)] = find_school_by_operator_suffix(value.get('operator'))
        if value.get('queue') == 'scholars-portal':
             ws['G'+str(max_row)] = user_show_value 
        elif value.get('queue') == 'scholars-portal-txt':
             ws['H'+str(max_row)]  = user_show_value 
        elif value.get('queue') == 'clavardez':
             ws['I'+str(max_row)] = user_show_value 
        elif value.get('queue') == 'practice-webinars':
            ws['J'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'practice-webinars-fr':
            ws['K'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'toronto-mississauga':
            ws['L'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'toronto-scarborough':
            ws['M'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'toronto-st-george':
            ws['N'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'brock':
            ws['O'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'carleton':
            ws['P'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'carleton-txt':
            ws['Q'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'laurentian':
            ws['R'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'laurentian-fr':
            ws['S'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'otech':
            ws['T'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'queens':
            ws['U'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'western':
            ws['V'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'western-txt':
            ws['W'+ str(max_row)] = user_show_value
        elif value.get('queue') == 'western-proactive':
            ws['X'+ str(max_row)] = user_show_value
        else:
            pass
    filename = get_filename()
    wb.save(filename=filename)
    #Insert to db

def find_assignements(assignments, user):
    availabilities = list()
    loc_dt = eastern.localize(datetime.now())
    staffing = False
    for assignment in assignments:
        if assignment['enabled']:
            staffing = True
            queue = assignment.get('queue')
            queueShow = assignment.get('queueShow')
            userShow = assignment.get('userShow')
            #print({'queue': queue, 'queueShow': queueShow, 'userShow':userShow})
            #insert to excel
            value = {
                'queue': queue,
                'operator': user.get('name'),
                'user_show': user.get('show'),
                'status': user.get('status'),
                'date': str(loc_dt.strftime(fmt_date)),
                'time':str(loc_dt.strftime(fmt_hour)), 
                'hour_floor': str(loc_dt.strftime(fmt_hour))[0:2],
                'hour': str(loc_dt.strftime(fmt_hour))
            }
            #print(value)
            availabilities.append(value)
    return [staffing, availabilities]


def webclient_activity():
    users = client.all('users')
    num_users = 0
    for user in users.get_list():
        if user.get('name') == 'guinsly_sp':
            #print(user) #breakpoint()
            pass
        if user['show'] == 'unavailable':
            continue

        # Is that user staffing any queue?
        assignments = users.one(user['id']).all('assignments').get_list()
        #print(assignments)
        result = find_assignements(assignments, user)
        
        staffing = result[0]
        availabilities = result[1]

        # insert to excel
        insert_to_excel(availabilities, user)


def is_hour_between(start, end, now):
    is_between = False

    is_between |= start <= now <= end
    is_between |= end < start and (start <= now or now <= end)

    return is_between

def find_which_weekday():
    day = datetime.today().weekday()
    if day >= 1 and day < 5:
        return [9.5, 22.15]
    elif day == 5:
        return [9.5, 17.15]
    else:
        #weekend
        return [11.5, 1.5]

if __name__ == '__main__':
    day_time = find_which_weekday()
    try_open_file()
    
    webclient_activity()

    # add cronjob for each minutes of the day
    # https://crontab.guru/between-certain-hours

    


