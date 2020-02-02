import requests 
import sched, time
import pytz
from pytz import timezone
utc = pytz.utc
eastern = timezone('America/Montreal')
from datetime import datetime

from openpyxl import Workbook, load_workbook

fmt_date = '%Y-%m-%d'
fmt_hour = '%H:%M:%S'
fmt_hour_round = '%H.%M'
from pprint import pprint as print

#availability CONSTANT
queues = ['scholars-portal', "scholars-portal-txt", "clavardez"]
start_url = "https://ca.libraryh3lp.com/presence/jid/" 
end_url =  "/chat.ca.libraryh3lp.com/text"

def get_filename():
    loc_dt = eastern.localize(datetime.now())
    filename = str(loc_dt.strftime(fmt_date)) + '.xlsx'
    return filename

def try_open_file():
    filename = get_filename()
    try:
        wb = load_workbook(filename)
        return wb
    except:
        wb = prepare_workbook()
        return wb

def insert_to_excel(availabilities):
    wb = try_open_file()
    ws = wb.active
    max_row = ws.max_row + 1
    for value in availabilities:
        ws['A'+str(max_row)] = value.get('date')
        ws['B'+str(max_row)] = value.get('time')
        ws['C'+str(max_row)] = value.get('hour_floor')
        if value.get('queue') == 'scholars-portal':
             ws['D'+str(max_row)] = value.get('response') 
        elif value.get('queue') == 'scholars-portal-txt':
             ws['E'+str(max_row)]  = value.get('response') 
        elif value.get('queue') == 'clavardez':
             ws['F'+str(max_row)] = value.get('response') 
        else:
            pass
    filename = get_filename()
    wb.save(filename=filename)
    #Insert to db

def find_availability_for_queues():
    availabilities = list()
    for queue in queues:
        response = requests.get(start_url+queue+end_url).content
        loc_dt = eastern.localize(datetime.now())
        value = {
            'queue': queue,
            'response': response.decode("utf-8") ,
            'date': str(loc_dt.strftime(fmt_date)),
            'time':str(loc_dt.strftime(fmt_hour)), 
            'hour_floor': str(loc_dt.strftime(fmt_hour))[0:2],
            'hour': str(loc_dt.strftime(fmt_hour))
        }
        #print(value)
        availabilities.append(value)
    return availabilities
    

def prepare_workbook():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'date'
    ws['B1'] = 'time'
    ws['C1'] = 'time_hour'
    ws['D1'] = 'scholars-portal'
    ws['E1'] = 'scholars-portal-txt'
    ws['F1'] = 'clavardez'
    loc_dt = eastern.localize(datetime.now())
    wb.save(str(loc_dt.strftime(fmt_date)) + '.xlsx')
    return wb

def is_hour_between(start, end, now):
    is_between = False

    is_between |= start <= now <= end
    is_between |= end < start and (start <= now or now <= end)

    return is_between

def find_which_weekday():
    day = datetime.today().weekday()
    if day >= 1 and day < 5:
        return [9.5, 22.15]
    elif day == 5:
        return [9.5, 17.15]
    else:
        #weekend
        return [11.5, 1.5]

if __name__ == '__main__':
    day_time = find_which_weekday()

    loc_dt = eastern.localize(datetime.now())
    time_now = (float(str(loc_dt.strftime(fmt_hour_round))))
    if (is_hour_between(day_time[0], day_time[1], time_now)):
        #if between 2 times
        availabilities = find_availability_for_queues()
        insert_to_excel(availabilities)
        print('in file')
    # add cronjob for each minutes of the day
    # https://crontab.guru/between-certain-hours

    

